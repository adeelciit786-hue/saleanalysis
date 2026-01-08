import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create a workbook
wb = Workbook()
ws = wb.active
ws.title = "December"

# Define styles
header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
total_fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
total_font = Font(bold=True, size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Row 1: Month name + Weekdays
month_name = "December"
weekdays = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]

ws['A1'] = month_name
ws['A1'].font = Font(bold=True, size=12)

for idx, day in enumerate(weekdays, start=2):
    cell = ws.cell(row=1, column=idx)
    cell.value = day
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Row 2: Date numbers (1-7 for first week)
dates = [1, 2, 3, 4, 5, 6, 7]
for idx, date in enumerate(dates, start=2):
    cell = ws.cell(row=2, column=idx)
    cell.value = date
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Add branch data (rows 3 onwards)
branches = [
    "AL BARRARI",
    "AL FORSAN",
    "DEIRA BRANCH",
    "BUR DUBAI",
    "JBR MALL",
    "IBN BATTUTA",
    "MARINA MALL",
    "CITY WALK",
    "GALLERIA",
    "MOE",
    "AL KHAWANEEJ",
    "AL WARQA",
    "MIRDIF",
    "AL KARAMA",
    "SATWA BRANCH",
    "MANARA CENTER",
    "AL FAHIDI",
    "JUMEIRAH",
    "AL SAFA",
    "BUSINESS BAY",
    "DOWNTOWN",
    "CREEK HARBOR",
    "PALM JUMEIRAH",
    "ARABIAN RANCHES",
    "MOTOR CITY",
    "SPORT CITY",
    "DISCOVERY GARDENS",
    "LIWAN"
]

# Sample sales data (random for demo)
import random
random.seed(42)

for branch_idx, branch in enumerate(branches, start=3):
    ws.cell(row=branch_idx, column=1).value = branch
    ws.cell(row=branch_idx, column=1).font = Font(bold=True)
    
    for day_idx in range(2, 9):  # 7 days
        cell = ws.cell(row=branch_idx, column=day_idx)
        # Generate realistic sales data
        base_sales = random.uniform(500, 2000)
        weekday = (day_idx - 2) % 7
        
        # Boost sales on weekends
        if weekday >= 5:
            base_sales *= 1.3
        
        cell.value = round(base_sales, 2)
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal='right')
        cell.border = border

# Add TOTAL row
total_row = len(branches) + 3
ws.cell(row=total_row, column=1).value = "TOTAL"
ws.cell(row=total_row, column=1).font = total_font
ws.cell(row=total_row, column=1).fill = total_fill

for day_idx in range(2, 9):
    cell = ws.cell(row=total_row, column=day_idx)
    # Sum formula for the column
    col_letter = get_column_letter(day_idx)
    cell.value = f"=SUM({col_letter}3:{col_letter}{total_row-1})"
    cell.font = total_font
    cell.fill = total_fill
    cell.number_format = '#,##0.00'
    cell.alignment = Alignment(horizontal='right')
    cell.border = border

# Adjust column widths
ws.column_dimensions['A'].width = 20
for col_idx in range(2, 9):
    ws.column_dimensions[get_column_letter(col_idx)].width = 15

# Save the workbook
wb.save('December_Sample.xlsx')
print("Sample Excel file created: December_Sample.xlsx")
